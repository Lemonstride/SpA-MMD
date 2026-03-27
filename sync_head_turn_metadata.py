from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any


DEFAULT_ROOT_DIR = Path("/your/own/path/processed")
DEFAULT_XLSX_PATH = Path("/your/own/path/total.xlsx")
TARGET_METRIC_HEADERS = [
    "left_max_angle",
    "left_max_angle_deg",
    "right_max_angle",
    "right_max_angle_deg",
    "total_rom_deg",
    "asymmetry_deg",
]
BASE_LABEL_HEADERS = [
    "subject_id",
    "name",
    "gender",
    "age",
    "height",
    "weight",
    "test_time",
]
PATIENT_ONLY_LABEL_HEADERS = [
    "course_of_disease",
    "main_diagnosis",
    "walking_ability",
    "limp_or_not",
    "stable_ability",
]
HEADER_ALIASES = {
    "subject_id": "subject_id",
    "姓名": "name",
    "name": "name",
    "性别": "gender",
    "gender": "gender",
    "年龄": "age",
    "age": "age",
    "身高": "height",
    "height": "height",
    "体重": "weight",
    "weight": "weight",
    "测试时间": "test_time",
    "test_time": "test_time",
    "病程": "course_of_disease",
    "course_of_disease": "course_of_disease",
    "主要诊断": "main_diagnosis",
    "main_diagnosis": "main_diagnosis",
    "步行能力等级": "walking_ability",
    "walking_ability": "walking_ability",
    "是否跛行": "limp_or_not",
    "limp_or_not": "limp_or_not",
    "平衡能力": "stable_ability",
    "stable_ability": "stable_ability",
    "severity": "severity",
    "left_max_angle": "left_max_angle",
    "left_max_angle_deg": "left_max_angle_deg",
    "right_max_angle": "right_max_angle",
    "right_max_angle_deg": "right_max_angle_deg",
    "total_rom_deg": "total_rom_deg",
    "asymmetry_deg": "asymmetry_deg",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sync patient metadata and head-turn summary metrics between processed sessions and total.xlsx."
    )
    parser.add_argument("--root-dir", type=Path, default=DEFAULT_ROOT_DIR, help="Processed dataset root.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX_PATH, help="Workbook path such as total.xlsx.")
    parser.add_argument("--sheet-name", type=str, default=None, help="Optional worksheet name. Defaults to active sheet.")
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        help="Optional output workbook path. Defaults to overwriting the input workbook.",
    )
    return parser.parse_args()


def load_workbook_or_exit(path: Path) -> tuple[Any, Any]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "缺少 openpyxl。请先安装后再运行:\n"
            "python -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(path)
    return load_workbook, workbook


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def header_map(worksheet: Any) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, cell in enumerate(worksheet[1], start=1):
        header = normalize_header(cell.value)
        if header:
            mapping[header] = index
    return mapping


def ensure_header(worksheet: Any, headers: dict[str, int], header_name: str) -> int:
    if header_name in headers:
        return headers[header_name]
    new_index = len(headers) + 1
    worksheet.cell(row=1, column=new_index, value=header_name)
    headers[header_name] = new_index
    return new_index


def read_row_dict(worksheet: Any, row_index: int, headers: dict[str, int]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for header, column_index in headers.items():
        row[header] = worksheet.cell(row=row_index, column=column_index).value
    return row


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)


def read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def clean_patient_metadata(row: dict[str, Any]) -> dict[str, Any]:
    excluded = {
        "",
        "left_max_angle",
        "left_max_angle_deg",
        "right_max_angle",
        "right_max_angle_deg",
        "total_rom_deg",
        "asymmetry_deg",
    }
    metadata: dict[str, Any] = {}
    for key, value in row.items():
        canonical_key = HEADER_ALIASES.get(key, key)
        if canonical_key in excluded:
            continue
        if value is None:
            continue
        metadata[canonical_key] = value
    return metadata


def update_json_with_metadata(path: Path, patient_metadata: dict[str, Any]) -> None:
    payload: dict[str, Any] = {}
    if path.exists():
        payload = read_json(path)
    payload["clinical_metadata"] = patient_metadata
    severity = patient_metadata.get("severity")
    if severity is not None:
        payload["severity_label"] = severity
        payload["binary_label"] = 0 if float(severity) == 0 else 1
    write_json(path, payload)


def format_label_value(value: Any) -> str:
    if value is None:
        return "unknown"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def build_disease_annotations(patient_metadata: dict[str, Any], existing_payload: dict[str, Any] | None = None) -> dict[str, Any]:
    severity = patient_metadata.get("severity")
    binary_label = None if severity is None else (0 if float(severity) == 0 else 1)

    payload: dict[str, Any] = {
        "binary_label": binary_label,
        "severity_label": severity,
        "notes": "",
    }
    if existing_payload is not None:
        payload["notes"] = existing_payload.get("notes", "")

    for key in BASE_LABEL_HEADERS:
        value = patient_metadata.get(key)
        if value is not None:
            payload[key] = value

    if binary_label == 1:
        for key in PATIENT_ONLY_LABEL_HEADERS:
            value = patient_metadata.get(key)
            if value not in (None, ""):
                payload[key] = value

    return payload


def update_label_files(label_dir: Path, patient_metadata: dict[str, Any]) -> int:
    label_dir.mkdir(parents=True, exist_ok=True)
    existing_annotations: dict[str, Any] | None = None
    disease_annotations_path = label_dir / "disease_annotations.json"
    if disease_annotations_path.exists():
        existing_annotations = read_json(disease_annotations_path)

    annotations_payload = build_disease_annotations(patient_metadata, existing_annotations)
    write_json(disease_annotations_path, annotations_payload)

    binary_value = annotations_payload.get("binary_label")
    severity_value = annotations_payload.get("severity_label")

    (label_dir / "binary_label.txt").write_text(
        format_label_value(binary_value),
        encoding="utf-8",
    )
    (label_dir / "severity_label.txt").write_text(
        format_label_value(severity_value),
        encoding="utf-8",
    )
    return 3


def main() -> None:
    args = parse_args()
    root_dir = args.root_dir.expanduser().resolve()
    xlsx_path = args.xlsx.expanduser().resolve()
    output_xlsx = args.output_xlsx.expanduser().resolve() if args.output_xlsx is not None else xlsx_path

    if not root_dir.exists():
        raise SystemExit(f"找不到 root-dir: {root_dir}")
    if not xlsx_path.exists():
        raise SystemExit(f"找不到 xlsx: {xlsx_path}")

    _, workbook = load_workbook_or_exit(xlsx_path)
    worksheet = workbook[args.sheet_name] if args.sheet_name else workbook.active
    headers = header_map(worksheet)

    if "subject_id" not in headers:
        raise SystemExit("表格缺少 subject_id 列")

    metric_columns = {
        "left_max_angle_deg": ensure_header(worksheet, headers, "left_max_angle"),
        "right_max_angle_deg": ensure_header(worksheet, headers, "right_max_angle"),
        "total_rom_deg": ensure_header(worksheet, headers, "total_rom_deg"),
        "asymmetry_deg": ensure_header(worksheet, headers, "asymmetry_deg"),
    }

    processed_subjects = 0
    processed_sessions = 0
    updated_json_files = 0
    updated_excel_rows = 0

    row_index = 2
    while True:
        subject_id = worksheet.cell(row=row_index, column=headers["subject_id"]).value
        if subject_id in (None, ""):
            break

        subject_id_str = str(subject_id).strip()
        row = read_row_dict(worksheet, row_index, headers)
        patient_metadata = clean_patient_metadata(row)

        subject_dir = root_dir / subject_id_str

        for session_name in ("head_turn", "walk"):
            session_dir = subject_dir / session_name
            if not session_dir.exists():
                continue

            labels_dir = session_dir / "labels"
            updated_json_files += update_label_files(labels_dir, patient_metadata)

            summary_path = labels_dir / "head_turn_state" / "summary.json"
            if session_name == "head_turn" and summary_path.exists():
                summary_payload = read_json(summary_path)
                summary_payload["clinical_metadata"] = patient_metadata
                if "severity" in patient_metadata:
                    summary_payload["severity_label"] = patient_metadata["severity"]
                    summary_payload["binary_label"] = 0 if float(patient_metadata["severity"]) == 0 else 1
                write_json(summary_path, summary_payload)
                updated_json_files += 1

                for metric_key, column_index in metric_columns.items():
                    worksheet.cell(row=row_index, column=column_index, value=summary_payload.get(metric_key))
                updated_excel_rows += 1

            for json_path in (session_dir / "meta.json", session_dir / "session_meta.json"):
                if json_path.exists():
                    update_json_with_metadata(json_path, patient_metadata)
                    updated_json_files += 1

            processed_sessions += 1

        processed_subjects += 1
        row_index += 1

    workbook.save(output_xlsx)

    print(f"处理受试者数: {processed_subjects}")
    print(f"处理 session 数: {processed_sessions}")
    print(f"更新 JSON 文件数: {updated_json_files}")
    print(f"更新 Excel 行数: {updated_excel_rows}")
    print(f"输出 Excel: {output_xlsx}")


if __name__ == "__main__":
    main()
